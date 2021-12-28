import openpyxl as xl
from openpyxl.chart import BarChart, Reference

## open required spreadsheet
wb = xl.load_workbook('transactions.xlsx')

## access specific sheet from spreadsheet
sheet = wb['Sheet1']

## loop over the all of rows in sheet
for row in range(2, sheet.max_row + 1):
  # 1. get the cell for each row
  cell = sheet.cell(row, 3)

  # 2. modify cell value as per req
  corrected_price = cell.value * 0.9

  # 3. reference/create a new cell for each row
  corrected_price_cell = sheet.cell(row, 4)

  # 4. set the value for these cells
  corrected_price_cell.value = corrected_price

## select the range for charts
values = Reference(
sheet, # which sheet to work with
min_row = 2, # start from second row
max_row = 4, # go till 4th row
min_col = 4, # start with 4th col
max_col = 4 # go till 4th col
)

## create a bar chart
chart = BarChart()
# add data to chart
chart.add_data(values)

## add chart to sheet
sheet.add_chart(chart, 'e2') # `e2` is the location of chart on the sheet

## save sheet in a new file
wb.save('transactions2.xlsx')
